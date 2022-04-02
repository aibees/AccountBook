from datetime import datetime

import openpyxl

from common.database.Connect import Connection
from common.database.sql.SlipQuery import InsertSlip


def saveInsertData(data):
    result = '000'
    conn = Connection()
    try:
        for row in data:
            conn.insert(InsertSlip, row)
    except Exception as e:
        print(e)
        result = e.__cause__
    return result


def excelDataParse(fileName):
    result = []
    workBook = openpyxl.load_workbook(filename=fileName).active  # 첫번째 sheet 가져오기
    for r in range(2, workBook.max_row + 1):
        c_r = str(r)
        row = []
        for c in range(workBook.max_column - 1):
            row.append(workBook[chr(ord('A') + c) + c_r].value)

        #record time
        date = workBook['A' + c_r].value
        recode = workBook['L' + c_r].value
        record_time = date[:4] + '-' + date[4:6] + '-' + date[-2:] + 'T' + recode[:2] + ':' + recode[2:4] + ':' + recode[-2:]
        row.append(record_time)
        result.append(row)
    return result
