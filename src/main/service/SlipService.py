from datetime import datetime

import openpyxl

from common.database.Connect import Connection
from common.database.sql.SlipQuery import InsertSlip


def saveInsertData(data):
    result = '000'
    conn = Connection()
    try:
        for row in data:
            conn.insert(InsertSlip, row)
    except Exception as e:
        print(e)
        result = e.__cause__
    return result


def excelDataParse(fileName):
    result = []
    workBook = openpyxl.load_workbook(filename=fileName).active  # 첫번째 sheet 가져오기
    for r in range(2, workBook.max_row + 1):
        c_r = str(r)
        row = []
        for c in range(workBook.max_column):
            cell = workBook[chr(ord('A') + c) + c_r].value
            if isinstance(cell, datetime):
                row.append(cell.strftime("%Y%m%d"))
            else:
                row.append(workBook[chr(ord('A') + c) + c_r].value)
        # recordtime 시간은 엑셀 넣은 시간으로, 날짜는 엑셀 안에 기록된 날짜로
        row.append(workBook[chr(ord('A')) + c_r].value.strftime("%Y-%m-%d") + 'T' + datetime.now().strftime("%H:%M:%S"))
        result.append(row)
    return result
